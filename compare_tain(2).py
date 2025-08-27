# -*- coding: utf-8 -*-



"""compare_tain.ipynb
### Libraries
- Import required libraries for creating and analysing the networks
"""

!pip install graphkit-learn #for GED

import os #for alt names xlsx
from openpyxl import load_workbook #for alt names xlsx
import numpy as np
import networkx as nx
import matplotlib.pyplot as plt
from prettytable import PrettyTable #for printing nicer looking tables
import scipy
from scipy import stats
from scipy import integrate
import random

#GED
from gklearn.ged.env import GEDEnv

#NMI
from scipy.special import loggamma
from collections import defaultdict

#MLE
from scipy.optimize import minimize
from scipy.special import zeta
from scipy.special import factorial
from scipy.stats import poisson

"""### Data
- Import social network data
"""

# Datasets *excluding* large groups
kinsella = 'Kinsella.txt'
kinsella_np = 'Kinsella_NoPretales.txt'
orahilly1e = 'ORahilly_R1_English_GC.xlsx'
orahilly1i = 'ORahilly_R1_Irish.xlsx'
orahilly2 = 'ORahilly_R2_Irish.xlsx'
carson = 'Carson.txt'
nogroups = [orahilly1e,
            orahilly1i,
            orahilly2,
            kinsella,
            kinsella_np,
            carson]

# Datasets *including* large groups
kinsella_g = 'Kinsella_Groups.txt'
kinsella_np_g = 'Kinsella_NoPretales_Groups.txt'
orahilly1_ge = 'ORahilly_R1_English_Groups_GC.xlsx'
orahilly1_gi = 'ORahilly_R1_Irish_Groups.xlsx'
orahilly2_g = 'ORahilly_R2_Irish_Groups.xlsx'
carson_g = 'Carson_Groups.txt'
groups = [orahilly1_ge,
          orahilly1_gi,
          orahilly2_g,
          kinsella_g,
          kinsella_np_g,
          carson_g
          ]




"""### Network Creation
- Create social networks using the data imported above
"""

def create_network(input_data):
  '''
  === Summary ===
  Creates a network from a text file (.txt) or Excel (.xlsx) file containing character interactions, maps alternate names to appropriate character names based on the 'Characters list_GC.xlsx' file.

  === Inputs ===
  input_data : The path to the input file containing character interaction data (.txt or .xlsx).

  === Outputs ===
  G : An undirected weighted graph representing the character interactions in the input data.
    - Nodes represent characters
    - Edges represent interactions between characters
    - Edge weights indicate frequency of interactions
  '''
  alts = {} # Dictionary for mapping alternate names to 'main' names
  chars = [] # List of all 'main' character names
  check = [] # List to track potential duplicate names

  chars_data = load_workbook('Characters list_GC.xlsx') # Loads alternate names from reference Excel file
  data = chars_data.active # Active workbook
  headers = [cell.value for cell in data[1]] # Isolate headers

  # Iterate through names from reference Excel file
  for i in range(2,data.max_row+1):
    u = data[i][0].value
    if u is not None and u[0] == '%':
        continue
    u = u.strip()
    chars.append(u)
    # Find the alternate names for this character
    l = [cell.value for cell in data[i][1:data.max_column] if cell.value]
    for a in l:
        a = a.strip()
        if len(a) > 0:
            if a in alts:
                check.append(a)
            alts[a] = u

  G = nx.Graph() # Initialise empty grah


  # === ------------------------------ ===
  # === Create network for .xlsx files ===
  # === ------------------------------ ===
  if input_data.split('.')[1] == 'xlsx':
    data = load_workbook(input_data)
    data = data.active
    headers = [cell.value for cell in data[1]]

    # Idenitfy columns for required fields
    fr = headers.index('Friendly')
    ho = headers.index('Hostile')
    if 'Familial' in headers:
      fa = headers.index('Familial')
    else:
      None

    input_data = input_data.split('.')[0] # Remove file extention

    ## Append nodes and edges to network
    for i in range(2,data.max_row+1):
      u = data[i][0].value
      if u != None and u[0] != '%':
        u = u.strip()

        # Resolve the alternate name if necessary
        if u in alts:
          u = alts[u]

        # Isolate friendly/hostile/familial edges
        friendly_links = [cell.value for cell in data[i][fr:ho] if cell.value]
        if 'Familial' in headers:
          hostile_links = [cell.value for cell in data[i][ho:fa] if cell.value]
          familial_links = [cell.value for cell in data[i][fa:] if cell.value]
        else:
          hostile_links = [cell.value for cell in data[i][ho:] if cell.value]
          familial_links = []

        # Add friendly links to network
        for v in friendly_links:
          v = v.strip()
          if v in alts:
            v = alts[v]
          if len(u) > 1 and len(v) > 1:
            if G.has_edge(u,v):
              G.get_edge_data(u,v)['weight'] += 1
            else:
              G.add_edge(u,v,weight=1,colour = 'green')

        # Add hostile links to network
        for v in hostile_links:
          v = v.strip()
          if v in alts:
            v = alts[v]
          if len(u) > 1 and len(v) > 1:
            if G.has_edge(u,v):
              G.get_edge_data(u,v)['weight'] += 1
            else:
              G.add_edge(u,v,weight=1,colour='red')

        # Add familial links to network (if present)
        if 'Familial' in headers:
          for v in familial_links:
            v = v.strip()
            if v in alts:
              v = alts[v]
            if len(u) > 1 and len(v) > 1:
              if G.has_edge(u,v):
                G.get_edge_data(u,v)['weight'] += 1
              else:
                G.add_edge(u,v,weight=1,colour='green') # colour set to green, familial edge = friendly edge

  # === ----------------------------- ===
  # === Create network for .txt files ===
  # === ----------------------------- ===
  else:
    with open(input_data,'rb') as f:
      headers = f.readline().decode().strip().split('\t')
      input_data = input_data.split('.')[0]

      # Idenitfy columns for required fields
      fr = headers.index('Friendly')
      ge = headers.index('Gender')
      ho = headers.index('Hostile')
      if 'Familial' in headers:
        fa = headers.index('Familial')
      else:
        None

      # Iterate through each line, adding nodes/edges to network
      for line in f:
        l = line.decode('latin-1').encode("utf-8").decode().strip().split('\t')
        if len(l) > 1:
          u = l[0]

          # Check for alternate names
          if u != None and u[0] != '%':
            u = u.strip()
            if u in alts:
              u = alts[u]

            # Isolate friendly/hostile/familial edges
            friendly_links = [v.strip() for v in l[fr:ho]]
            if 'Familial' in headers:
              hostile_links = [v.strip() for v in l[ho:fa]]
              familial_links = [v.strip() for v in l[fa:]]
            else:
              hostile_links = [v.strip() for v in l[ho:]]
              familial_links = []

            # Add friendly links to network
            for v in friendly_links:
              v = v.strip()
              if v in alts:
                v = alts[v]
              if len(u) > 1 and len(v) > 1:
                if G.has_edge(u,v):
                  G.get_edge_data(u,v)['weight'] += 1
                else:
                  G.add_edge(u,v,weight=1,colour='green')

            # Add hostile links to network
            for v in hostile_links:
              v = v.strip()
              if v in alts:
                v = alts[v]
              if len(u) > 1 and len(v) > 1:
                if G.has_edge(u,v):
                  G.get_edge_data(u,v)['weight'] += 1
                else:
                  G.add_edge(u,v,weight=1,colour='red')

            # Add familial links to network (if present)
            if 'Familial' in headers:
              for v in familial_links:
                v = v.strip()
                if v in alts:
                  v = alts[v]
                if len(u) > 1 and len(v) > 1:
                  if G.has_edge(u,v):
                    G.get_edge_data(u,v)['weight'] += 1
                  else:
                    G.add_edge(u,v,weight=1,colour='green') # colour set to green, familial edge = friendly edge

  G.remove_edges_from(nx.selfloop_edges(G))# Remove self-loops

  return G




"""### Network Quantities
- Calculate network structural properties
"""

def network_quantities(G,table=True):
  '''
  === Summary ===
  Calculates structural properties of a network.

  === Inputs ===
  G: An undirected NetworkX graph.
  table: Optional, if True, prints result in a PrettyTable, if False, results returned without printing a table

  === Outputs ===
  (N,L,k_avg,k_delta,r,C,C_rand,l_avg,l_rand)
  N: Number of nodes in the network
  L: Number of edges in the network
  k_avg: Average degree of the network
  k_delta: Degree delta of the network
  r: Assortativity coefficient of the network
  C: Clustering coefficient of the network
  C_rand: Clustering coefficient of a randomly generated network of the same size
  l_avg: Average path length of the network
  l_rand: Average path length of a randomly generated network of the same size
  '''
  N = G.number_of_nodes() # Number of nodes, N

  L = G.number_of_edges() # Number of edges, L

  degs = []
  for (node,value) in G.degree():
    degs.append(value)
  totdeg = sum(degs)
  k_avg = totdeg/N # Average degree, <k>

  degdiff = []
  for i in degs:
    degdiff.append((i-k_avg)**2)
  k_delta = sum(degdiff)/N # Degree variance, <k^2>

  r = nx.degree_assortativity_coefficient(G) # Assortativity, r

  C = nx.average_clustering(G) # Clustering coefficient, C

  if nx.is_connected(G) == False:
    avgpaths = []
    for x in nx.connected_components(G):
      subgraphs = G.subgraph(x)
      avgpaths.append(nx.average_shortest_path_length(subgraphs))
      l_avg = max(avgpaths) # Average path length, l_avg
  else:
    l_avg = nx.average_shortest_path_length(G) # Average path length, l_avg

  EM = 0.5772 # Euler-Macheroni constant
  lrand = 0.5 + (np.log(N)-EM)/np.log(k_avg) # Average path length of a random network of the same size, l_rand
  Crand = k_avg/(N-1) # Clustering coefficient of a randomly generated network of the same size C_rand

  # If true, table of network global properties is printed
  if table:
    network_properties = PrettyTable()
    network_properties.field_names = ['Property', 'Value']
    network_properties.add_row(['N', N])
    network_properties.add_row(['L', L])
    network_properties.add_row(['<k>', round(k_avg,2)])
    network_properties.add_row(['<k^2>', round(k_delta,2)])
    network_properties.add_row(['r', round(r,2)])
    network_properties.add_row(['C', round(C,2)])
    network_properties.add_row(['C_rand', round(Crand, 2)])
    network_properties.add_row(['l_avg', round(l_avg,2)])
    network_properties.add_row(['l_rand', round(lrand,2)])
    print(network_properties)
  else:
    None

  return N,L,k_avg,k_delta,r,C,Crand,l_avg,lrand




"""### Plot Network
- Create visualisation of the networks
"""
def plot_network(G,save_path=None):
  '''
  === Summary ===
  Creates a plot of a NetworkX graph.

  === Inputs ===
  G: An undirected NetworkX graph.

  === Outputs ===
  Plots the graph using Matplotlib.
  '''
  # Force plot to be printed as an output
#   %matplotlib inline

  # Layout
  pos = nx.kamada_kawai_layout(G)

  # Get top nodes by degree
  top_nodes = sorted(G.degree, key=lambda x: x[1], reverse=True)[:15]
  top_nodes_set = {node for node, x in top_nodes}

  # Print labels only top ranked nodes in terms of degree
  labels = {node: node if node in top_nodes_set else '' for node in G.nodes()}

  # Edge colours and opacity based on the weight of the edge
  edge_weights = [G[u][v]["weight"] for u, v in G.edges()]
  max_w = max(edge_weights) if edge_weights else 1
  # Scale opacity relative to the weight
  edge_alphas = []
  for i in edge_weights:
    edge_alphas.append(i/max_w)
  edge_colours = [G[u][v].get("colour") for u, v in G.edges()]

  # Split nodes to allow for top ranked nodes to be added on top of the rest
  topnodes = list(top_nodes_set)
  othernodes = []
  # Iterate through the nodes in G
  for i in G.nodes():
    # If the node is not in the list of top nodes, add it to the other set
    if i not in topnodes:
      othernodes.append(i)

  # Plot
  plt.figure(figsize=(18,10))
  plt.axis('off')

  # Define degree dictionary to make node sizes proportional to their degree
  degree_dict = dict(G.degree())

  # Other nodes added first
  nx.draw_networkx_nodes(G,pos,nodelist=othernodes,
                        node_size=[100 + degree_dict[n] * 4.5 for n in othernodes],
                        node_color= '#c0f1fc',
                        edgecolors='black',
                        linewidths= 0.1)

  # Top 10 nodes in terms of degree added next
  nx.draw_networkx_nodes(G,pos,nodelist=topnodes,
                        node_size=[100 + degree_dict[n] * 4.5 for n in topnodes],
                        node_color= '#c5f3c3',
                        edgecolors='#80e47c',
                        linewidths= 1)

  # Adds edges
  nx.draw_networkx_edges(G,pos,edge_color=edge_colours,
                         width=1,alpha=edge_alphas)

  # Adds labels
  nx.draw_networkx_labels(G, pos, labels=labels, font_size=7, font_weight='bold')

  if save_path:
    plt.savefig(save_path, dpi=800)

  plt.show()

  return





"""### Graph Matching
- Ipsen-Mikhailov Distance
- Graph Edit Distance
- Network Mutual Information
"""

def graph_matching(G1,G2):
  '''
  === Summary ===
  Calculates the Ipsen-Mikhailov Distance, Graph Edit Distance and the
  Network Mutual Information values between two networks.

  === Inputs ===
  G1, G2 : Networks to compare

  === Outputs ===
  im,ged,nmi : Values for the three graph matching methods
  '''

  # ===--------------------------===
  # === Ipsen-Mikhailov Distance ===
  # ===--------------------------===
  def gamma(G1,G2,n_decimal=2,lower=0.1,upper=0.7):
    """
    === Summary ===
    gamma of Lorentz distributions
    This function calculates the gamma parameter for a complete and empty network of
    the same size. Note the choice of the larger one to be empty to reduce computation time

    === Inputs ===
    G1, G2 : Networks to compare
    n_decimal : The number of decimal places we want gamma, it doesn't need to be particularly high
    lower :  The lower bound for where we want to begin looking for gamma
    upper : The upper bound for where we stop looking for an optimal gamma

    === Outputs ===
    gam : gamma of Lorentz distribution
    """
    n1, n2 = len(G1), len(G2)
    if n1 <= n2:
        F = nx.complete_graph(n1)
        E = nx.empty_graph(n2)
    else:
        F = nx.complete_graph(n2)
        E = nx.empty_graph(n1)
    s = '.'
    for i in range(n_decimal):
        s += '9'
    for gam in range(int(lower*(10**n_decimal)),int(upper*(10**n_decimal))):
        gam = gam/(10**n_decimal)
        d = IM_dist(F,E,gam)
        if round(d,n_decimal) == float(s):
            break
    return gam

  def IM_dist(G1,G2,gamma,weight=None):
    """
    === Summary ===
    Ipsen-Mikhailov distance
    This function first takes the Laplacian spectrum (i.e. the eignevectors of the Laplacian)
    The distance is obtained from the intergal of the difference of the Lorentz functions.

    === Parameters ===
    G1, G2 : Networks to compare
    gamma : The parameter gamma for the Lorentz function (see following function)
    weight : Name of the networkx weight property if assigned to the edges

    === Returns ===
    IM distance : The Ipsen-Mikhailov distance between the two graphs
    """
    L_g = nx.laplacian_spectrum(G1, weight=weight)
    L_h = nx.laplacian_spectrum(G2, weight=weight)
    L_g[L_g< 1e-10]=0
    L_h[L_h< 1e-10]=0
    omega_g = np.sqrt(L_g)
    omega_h = np.sqrt(L_h)

    lorentz = lambda w,gamma,omega: (gamma / ((w-omega)**2 + gamma**2)).sum()

    #These are the normalisation values for the Lorentz function of each graph
    K_g,err = integrate.quad(lorentz,0,np.inf,args=(gamma,omega_g,), limit=500)
    K_h,err = integrate.quad(lorentz,0,np.inf,args=(gamma,omega_h,), limit=500)
    K_g = 1/K_g
    K_h = 1/K_h

    d = lambda w: (K_g*lorentz(w,gamma,omega_g) - K_h*lorentz(w,gamma,omega_h))**2

    dist,err = integrate.quad(d,0,np.inf,limit=500)

    return round(np.sqrt(dist),2)

  # Find gamma, then the IM distance value between the two networks
  gam = gamma(G1,G2,n_decimal=1)

  im = IM_dist(G1,G2,gam) # IM


  # ===---------------------===
  # === Graph Edit Distance ===
  # ===---------------------===
  ged_env = GEDEnv() # initailize GED environment.
  ged_env.set_edit_cost('CONSTANT', # GED cost type.
                          edit_cost_constants=[1, 1, 1, 1, 1, 1] # edit costs.
                          )

  ged_env.add_nx_graph(G1, '') # add graph1
  ged_env.add_nx_graph(G2, '') # add graph2
  listID = ged_env.get_all_graph_ids() # get list IDs of graphs
  ged_env.init(init_type='LAZY_WITH_SHUFFLED_COPIES') # initialize GED environment.
  options = {'initialization_method': 'RANDOM', # or 'NODE', etc.
              'threads': 1 # parallel threads.
              }
  ged_env.set_method('BIPARTITE', # GED method.
                      options # options for GED method.
                      )
  ged_env.init_method() # initialize GED method.

  ged_env.run_method(listID[0], listID[1]) # run.

  ged = ged_env.get_upper_bound(listID[0], listID[1])	# GED


  # ===----------------------------===
  # === Network Mutual Information ===
  # ===----------------------------===

  # Helcio Felippe
  # https://github.com/hfelippe/network-MI
  # https://www.nature.com/articles/s42005-024-01830-3
  def zero_log(x):
    """log of zero is zero"""
    if x <= 0: return 0
    else: return np.log(x)
  def ent(vec):
    """entropy of a distribution"""
    vec  = np.array(vec)/sum(vec)
    return -sum([x*zero_log(x) for x in vec])
  def NMI(N,e1,e2):
    """normalized mutual information between N-node graphs with edge sets e1, e2"""
    Nc2 = N*(N-1)/2
    E1,E2,E12,Union = len(e1),len(e2),len(e1.intersection(e2)),len(e1.union(e2))
    p1,p2,p12 = E1/Nc2,E2/Nc2,E12/Nc2
    H1,H2 = ent([p1,1-p1]), ent([p2,1-p2])
    MI = H1 + H2 - ent([p12,p1-p12,p2-p12,1-p1-p2+p12])
    NMI = (2*MI+1e-100)/(H1+H2+1e-100) # negligibly small constants for the empty and complete graphs
    return NMI
  def add_nodes(g1,g2):
    extra1 = [u for u in g1 if u not in g2]
    extra2 = [u for u in g2 if u not in g1]
    g1.add_nodes_from(extra2)
    g2.add_nodes_from(extra1)
    return(g1,g2)

  # Ensure both graphs have same node set
  g1, g2 = add_nodes(G1.copy(), G2.copy())

  nmi = round(NMI(len(g1), set(g1.edges()), set(g2.edges())),2) # NMI

  return im,ged,nmi




"""### Print Comparison Matrix Tables
- This function creates a matrix of values for the different network comparisons methods
"""

def matrix_table(matrix,name,file_list):
  '''
  === Summary ===
  Prints matrix tables of distance measures

  === Inputs ===
  matrix : dictionary-of-dictionaries with the pairwise results.
  name : the metric being printed (KS test, GED, etc.).
  file_list : list of all files being compared

  === Outputs ===
  Prints table of distance measures
  '''

  # Print the name of the metric being printed
  print(f"\n{name}:")

  # Create an empty pretty table object to be filled with values
  table = PrettyTable()

  # Add column headers, first column is empty
  table.field_names = [""] + file_list

  # Iterate through each file in the list of files
  for f1 in file_list:
    row = [f1]

    # Iterate through the same list of files again,adding the values to the matrix
    for f2 in file_list:
      val = matrix[f1].get(f2)

      # Round the continuous values to two decimal places
      if name in ['Ipsen-Mikhailov Distance', 'Normalized Mutual Information','Kolmogorov-Smirnov Test','Spearmans Rank','Jaccard Index']:
        row.append(f"{val:.2f}")

      # Don't need to round for GED since the values are integers
      elif name == 'Graph Edit Distance':
        row.append(f"{int(val)}")

    # Add the row with the values to the table
    table.add_row(row)

  print(table)




"""### Degree Distributions
- Maximum Likelihood Estimates
- Forced Power Law Comparison
- Kolmogorov-Smirnov Tests
"""

def degree_distributions(G,kmin):

  ## ===== MLE =====
  ### Shane Mannion
  #### https://academic.oup.com/comnet/article/11/4/cnad023/7227325
  #### https://github.com/Shaneul/MLE/blob/main/MLE_functions.py
  """
  Created on Tue Nov 15 09:46:30 2022

  @author: Shane Mannion
  Functions for fitting degree distributions to complex networks.
  """

  if MLE:
    def freqTable(G):
        if type(G) == nx.classes.graph.Graph:
            degree_dict = dict(G.degree())
            degree_list = list(degree_dict.values())
        else:
            degree_list = G
        degree_list.sort()
        unique_deg = []
        table = {}
        for n in degree_list:
            if n in table:
                table[n] += 1
            else:
                table[n] = 1
        for n in degree_list:
            if n not in unique_deg:
                unique_deg += [n]
        return np.array(degree_list), np.array(unique_deg), table

    def degree_list(G):
        if type(G) == nx.classes.graph.Graph:
            degree_dict = dict(G.degree())
            degree_list = list(degree_dict.values())
        else:
            degree_list = G
        degree_list.sort()
        return np.array(degree_list)

    def empirical(X_list):
        N,f = np.unique(X_list, return_counts=True)
        cumul = np.cumsum(f[::-1])[::-1]
        p = f/X_list.size
        P = cumul/X_list.size
        return N, P, p

    def CCDF(result,X, N, P):
        k_min = result[0]
        Input = np.unique(X)
        distribution = result[1]
        params = result[2][0]
        C_index = np.where(N == k_min)[0]
        C = P[C_index]
        try:
            inf = np.arange(1000)#np.arange(np.amax(Input))
        except ValueError:
            inf = np.arange(1000)
        if distribution == 'Powerlaw':
            y = C*zeta(params[0], Input)/zeta(params[0], k_min)
        if distribution == 'Exponential':
            y = C*np.exp((-1/params[0])*(Input-k_min))
        if distribution == 'Weibull':
            sum1 = np.array([np.sum((((j+inf)/params[0])**(params[1]-1))*np.exp(-(((j+inf)/params[0])**params[1]))) for j in Input])
            inf_sum = np.sum((((inf + k_min)/params[0])**(params[1]-1))*np.exp(-1*((inf + k_min)/params[0])**params[1]))
            y = C*sum1/inf_sum
        if distribution == 'Lognormal':
            sum1 = np.array([np.sum( (1.0/(j+inf))*np.exp(-((np.log(j+inf)-params[0])**2)/(2*(params[1]**2)))) for j in Input])
            inf_sum = np.sum( (1.0/(inf+k_min)) * np.exp(-((np.log(inf+k_min)-params[0])**2)/(2*params[1]**2) ) )
            y = C*sum1/(inf_sum)
        if distribution == 'Poisson':
            y = 1 - C*poisson.cdf(Input, params[0])
        if distribution == 'Trunc_pl':
            inf_sum = np.sum((inf + k_min)**(-1*params[1]) * np.exp(-1*inf/params[0]))
            z = np.array([np.sum((inf + i)**(-1*params[1]) * np.exp(-1*inf/params[0])) for i in Input])
            y = C*(np.exp(-(Input-k_min)/params[0]))*z/inf_sum
        if distribution == 'Normal':
            norm_n = np.sum( np.exp( -((inf-params[0])**2)/(2*params[1]**2) ))
            sum1 = np.array([np.sum(np.exp(-((j+inf-params[0])**2)/(2*params[1]**2))) for j in Input])
            y = C*sum1/norm_n
        return y

    def PDF(result, X, N, p):
        k_min = result[0]
        distribution = result[1]
        Input = np.unique(X)
        params = result[2][0]
        C_index = np.where(N == k_min)[0]
        C = p[C_index]
        try:
            inf = np.arange(np.amax(Input))
        except ValueError:
            inf = np.arange(1000)
        if distribution == 'Powerlaw':
            y = (C/zeta(params[0], k_min))*Input**-params[0]
        if distribution == 'Exponential':
            y = C*( (1-np.exp(-1/params[0]))/np.exp(-k_min/params[0]) )*np.exp(-Input/params[0])
        if distribution == 'Weibull':
            inf_sum = np.sum((((inf + k_min)/params[0])**(params[1]-1))*np.exp(-1*((inf + k_min)/params[0])**params[1]))
            y = C* ((Input/params[0])**(params[1]-1)) * (np.exp((-(Input/params[0])**params[1]))) / inf_sum
        if distribution == 'Lognormal': # Not done
          inf_sum = np.sum( (1.0/(inf+k_min)) * np.exp(-((np.log(inf+k_min)-params[0])**2)/(2*params[1]**2) ) )
          y = C* ( (1/Input) * np.exp(-((np.log(Input)-params[0])**2)/(2*params[1]**2)) ) / inf_sum
        if distribution == 'Poisson': # Not done
            y = 1 - C*poisson.pdf(Input, params[0])
        if distribution == 'Trunc_pl':
            y = C*( (np.exp(-k_min/params[0]))/zeta(params[0], k_min) ) * (Input**(-params[0])) * np.exp(-Input/params[0])
        if distribution == 'Normal':
            norm_n = np.sum( np.exp( -((inf-params[0])**2)/(2*params[1]**2) ))
            y = C* ( np.exp(-((Input-params[0])**2)/(2*params[1]**2)) ) / norm_n

        return y

    def CDF(result,X, N, P):
        k_min = result[0]
        Input = X
        distribution = result[1]
        params = result[2]
        C_index = np.where(N == k_min)[0]
        C = P[C_index]
        try:
            inf = np.arange(np.amax(Input))
        except ValueError:
            inf = np.arange(1000)
        if distribution == 'Powerlaw':
            y = C*zeta(params[0], Input)/zeta(params[0], k_min)
        if distribution == 'Exponential':
            y = C*np.exp((-1/params[0])*(Input-k_min))
        if distribution == 'Weibull':
            sum1 = np.array([np.sum((((j+inf)/params[0])**(params[1]-1))*np.exp(-(((j+inf)/params[0])**params[1]))) for j in Input])
            inf_sum = np.sum((((inf + k_min)/params[0])**(params[1]-1))*np.exp(-1*((inf + k_min)/params[0])**params[1]))
            y = C*sum1/inf_sum
        if distribution == 'Lognormal':
            sum1 = np.array([np.sum( (1.0/(j+inf))*np.exp(-((np.log(j+inf)-params[0])**2)/(2*(params[1]**2)))) for j in Input])
            inf_sum = np.sum( (1.0/(inf+k_min)) * np.exp(-((np.log(inf+k_min)-params[0])**2)/(2*params[1]**2) ) )
            y = C*sum1/(inf_sum)
        if distribution == 'Poisson':
            y = 1 - C*poisson.cdf(Input, params[0])
        if distribution == 'Trunc_pl':
            inf_sum = np.sum((inf + k_min)**(-1*params[1]) * np.exp(-1*inf/params[0]))
            z = np.array([np.sum((inf + i)**(-1*params[1]) * np.exp(-1*inf/params[0])) for i in Input])
            y = C*(np.exp(-(Input-k_min)/params[0]))*z/inf_sum
        if distribution == 'Normal':
            norm_n = np.sum( np.exp( -((inf-params[0])**2)/(2*params[1]**2) ))
            sum1 = np.array([np.sum(np.exp(-((j+inf-params[0])**2)/(2*params[1]**2))) for j in Input])
            y = C*sum1/norm_n
        return 1 - y

    def AIC(LnL: float, N:int, params:int = 1):
        if N < 4:
            AIC = -2*LnL + 2*params
        else:
            AIC = -2*LnL + 2*params + ((2*params*(params + 1)))/(N - params - 1)
        return AIC

    def BIC(LnL: float, N:int, params:int = 1):
        return params * np.log(N) - 2*LnL

    def powerlaw(params:np.ndarray, x:np.ndarray, sum_log, delta:float = 0, k_min:int = 1):
        NegLnL =  x.size*np.log(zeta(params[0], k_min)) + params[0]*(sum_log)
        return NegLnL

    def exp_dist(params:np.ndarray, x:np.ndarray, delta:float=0, k_min:int=1):
        NegLnL = -1 * x.size*(np.log(1-np.exp(-1/params[0]))) + (1/params[0])*(x.sum() - x.size*k_min)
        return NegLnL

    def weibull(params, x:np.ndarray, inf, sum_log, delta:float=0, k_min:int=1):
        inf_sum = np.sum((((inf + k_min)/params[0])**(params[1]-1))*np.exp(-1*((inf + k_min)/params[0])**params[1]))
        LnL = -x.size * np.log(inf_sum) - x.size * (params[1] - 1) * np.log(params[0])\
            + (params[1] - 1) * sum_log - np.sum((x/params[0])**params[1])
        NegLnL = -1 * LnL
        return NegLnL

    def normal(params, x, inf):
      norm_n = np.sum( np.exp( -((inf-params[0])**2)/(2*params[1]**2) ))
      NegLnL = x.size*np.log(norm_n) + np.sum(((x - params[0])**2)/(2*params[1]**2))
      return NegLnL

    def stretched_exp(params,x, inf, k_min):
      norm_s = np.sum( np.exp(-((k_min+inf)/params[0])**params[1] ))
      NegLnL = -1*( -x.size*np.log(norm_s) - np.sum((x/params[0])**params[1]))
      return NegLnL

    def trunc_powerlaw(params, x:np.ndarray, inf, delta:float, k_min:int=1):
        inf_sum = np.sum((inf + k_min)**(-1*params[1]) * np.exp(-1*inf/params[0]))
        LnL = x.size * np.log(1 - delta) + x.size * k_min/params[0] - x.size*np.log(inf_sum)\
            - (params[1]*np.log(x) + x/params[0]).sum()
        NegLnL = -1*LnL
        return NegLnL

    def logn(params, x, inf, sum_log, k_min=1):
        inf_sum = np.sum( (1.0/(inf+k_min)) * np.exp(-((np.log(inf+k_min)-params[0])**2)/(2*params[1]**2) ) )
        NegLnL = -1*( - x.size*np.log(inf_sum) - sum_log - np.sum( ((np.log(x)-params[0])**2)/(2*params[1]**2) ) )
        return NegLnL

    def poisson_dist(lam, x:np.ndarray, delta:float, k_min:int=1):
        m = np.arange(k_min)
        LnL = x.size * np.log(1 - delta) - np.log(1 - np.exp(-1*lam) * np.sum((lam**m)/factorial(m)))\
            - x.size * lam + np.log(lam) * x.sum() - np.sum(np.log(factorial(x)))
        NegLnL = -1*LnL
        return NegLnL

    def poisson_large_k(lam, x:np.ndarray):
        d1 = poisson.pmf(x, lam)
        d1 = d1[np.nonzero(d1)]
        NegLnL = -1 * np.sum(np.log(d1))
        return NegLnL


    def MLE(X:np.ndarray, k_min:int = 1, vt:int = 3, IC:str = 'AIC'):
        votes = [100,10,100,10,100] # array of numbers to create a standard deviation
                                    # greater than 0.1
        Results = {}
        Results['Powerlaw'] = {}
        Results['Exponential'] = {}
        Results['Weibull'] = {}
        Results['Normal'] = {}
        Results['Trunc_PL'] = {}
        Results['Lognormal'] = {}
        Results['Poisson'] = {}
      #  Results['Compound Poisson'] = {}
        stop = False
        while stop == False:#np.std(votes[-vt:]) >= 0.1: # while the last X votes have not been the same
                                          # where X is vt.
            x = X[X >= k_min] # only include degree values over kmin
            delta = (X[X < k_min].size/X.size) # fraction below kmin
            k_mean = x.mean() # mean degree for initial parameter guesses
            try:
                inf = np.arange(np.amax(x) + 1000) # list of numbers for infinite sums required below
            except ValueError:  #raised if x is empty.
                inf = 1000
            sum_log = np.sum(np.log(x))
            opt_pl = minimize(powerlaw, (2), (x, sum_log, delta, k_min), method = 'SLSQP', bounds = [(0.5, 4)])
            Results['Powerlaw'][k_min] = [opt_pl['x'], -1*opt_pl['fun']]
            opt_exp = minimize(exp_dist, (k_mean), (x, delta, k_min), method = 'SLSQP', bounds = ((0.5,k_mean + 20),))
            Results['Exponential'][k_min] = [opt_exp['x'], -1*opt_exp['fun']]
            opt_wb = minimize(weibull, (k_mean,1),(x, inf, sum_log, delta, k_min), method = 'SLSQP', bounds=((0.05, None),(0.05, 4),))
            Results['Weibull'][k_min] = [opt_wb['x'], - 1*opt_wb['fun']]
            opt_normal = minimize(normal, (k_mean, np.std(x)), (x, inf),method='SLSQP',bounds=[(0.,k_mean+10),(0.1,None)])
            Results['Normal'][k_min] = [opt_normal['x'], -1*opt_normal['fun']]
            opt_tpl = minimize(trunc_powerlaw,(k_mean,1),(x, inf, delta, k_min), method = 'SLSQP', bounds=((0.5, k_mean + 20),(0.5,4),))
            Results['Trunc_PL'][k_min] = [opt_tpl['x'], -1*opt_tpl['fun']]
            try: #prevents valueerror when value goes out of bounds given in function
                opt_logn = minimize(logn, (np.log(k_mean), np.log(x).std()), (x, inf, sum_log, k_min), method='TNC',bounds=[(0.,np.log(k_mean)+10),(0.01,np.log(x.std())+10)])
                Results['Lognormal'][k_min] = [opt_logn['x'], -1*opt_logn['fun']]
            except ValueError:
                Results['Lognormal'][k_min] = [[0,0], 10000]
            try:
                poisson_max = np.amax(x)
            except ValueError:
                poisson_max = 1
            if poisson_max > 170: #different method used when k_max is large, due to infinity from factorial
                opt_p = minimize(poisson_large_k, x.mean(), (x), method='SLSQP')
            else:
                opt_p = minimize(poisson_dist, x.mean(), (x, delta, k_min), method='SLSQP', bounds = ((0.5, None),))
            Results['Poisson'][k_min] = [opt_p['x'], -1*opt_p['fun']]
            Distributions = list(Results.keys())
            AICs = []
            BICs = []
            for i in Results.keys():
                if i == 'Lognormal':
                    if Results[i][k_min][0][1] == 0:
                        AICs.append(float("inf"))
                        BICs.append(float("inf"))
                if AIC(Results[i][k_min][1], x.size, len(Results[i][k_min][0])) == float("-inf"):
                    AICs.append(float("inf"))
                else:
                    AICs.append(AIC(Results[i][k_min][1], x.size, len(Results[i][k_min][0])))
                if BIC(Results[i][k_min][1], x.size, len(Results[i][k_min][0])) == float("-inf"):
                    BICs.append(float("inf"))
                else:
                    BICs.append(BIC(Results[i][k_min][1], x.size, len(Results[i][k_min][0])))
            weights = []
            weight_total = 0
            for i in AICs:
                weight_total += np.exp(-1*(i - np.min(AICs))/2)
            for i in AICs:
                weights += [np.exp(-1*(i - np.min(AICs))/2)/weight_total]
            if IC == 'AIC':
              votes.append(np.argmax(weights).astype(np.int32))
            if IC == 'BIC':
              votes.append(np.argmin(BICs).astype(np.int32))
            #if we only want to fit at a specific k_min, break the loop and return the first result
            if vt == 1:
                Delta = (X[X < k_min]).size/X.size
                Final_dist = [k_min, Distributions[np.argmax(weights)],Results[Distributions[np.argmax(weights)]][k_min], Delta]
                return Final_dist
            if vt > 1:
                if np.std(votes[-vt:]) <= 0.1:
                    stop = True
            k_min += 1
        Delta = (X[X < (k_min-vt)]).size/X.size
        Final_dist = [k_min-vt, Distributions[np.argmax(weights)],Results[Distributions[np.argmax(weights)]][k_min-vt], Delta]
        if len(weights) > 0:
            Final_dist.append(weights)
        return Final_dist

    def opt_single_dist(X, result, k_min):
        x = X[X >= k_min]
        delta = (X[X < k_min].size/X.size)
        k_mean = x.mean()
        try:
            inf = np.arange(np.amax(x)) #Creates a sequence of numbers for infinite sums
        except ValueError:  #raised if x is empty.
            inf = 1000
        sum_log = np.sum(np.log(x))
        if result[1] == 'Powerlaw':
            opt = minimize(powerlaw, (2), (x, sum_log, delta, k_min), method = 'SLSQP', bounds = [(0.5, 4)])
        if result[1] == 'Exponential':
            opt = minimize(exp_dist, (k_mean), (x, delta, k_min), method = 'SLSQP', bounds = ((0.5,k_mean + 20),))
        if result[1] == 'Weibull':
            opt = minimize(weibull, (k_mean,1),(x, inf, sum_log, delta, k_min), method = 'SLSQP', bounds=((0.05, None),(0.05, 4),))
        if result[1] == 'Normal':
            opt = minimize(normal, (k_mean, np.std(x)), (x, inf), method='SLSQP', bounds=[(0.,k_mean+10),(0.1,None)])
        if result[1] == 'Stretched_Exp':
            opt = minimize(stretched_exp,(k_mean,1),(x, inf, k_min), method='SLSQP',bounds=[(0.5,None),(0.05,4.)])
        if result[1] == 'Trunc_PL':
            opt = minimize(trunc_powerlaw,(k_mean,1),(x, inf, delta, k_min), method = 'SLSQP', bounds=((0.5, k_mean + 20),(0.5,4),))
        if result[1] == 'Lognormal':
            try: #prevents valueerror when value goes out of bounds given in function
                opt = minimize(logn, (np.log(k_mean), np.log(x).std()), (x, inf, sum_log, k_min),
                              method='TNC',bounds=[(0.,np.log(k_mean)+10),(0.01,np.log(x.std())+10)])
            except ValueError:
                return [0,0]
        if result[1] == 'Poisson':
            try:
                poisson_max = np.amax(x)
            except ValueError:
                poisson_max = 1
            if poisson_max > 170: #different method used when k_max is large, due to infinity from factorial
                opt = minimize(poisson_large_k, x.mean(), (x), method='SLSQP')
            else:
                opt = minimize(poisson_dist, x.mean(), (x, delta, k_min), method='SLSQP', bounds = ((0.5, None),))
        return opt['x']

    #Plotting function has been altered to generate a single output plot which contains all of the data
    def plotting(N, Input, fit, result, emp, dist, Name = '', save=False, saveloc=''):
      dataset_styles = {# ORahilly group
      "ORahilly_R1_English_GC.xlsx": {"colour": "#029d3f", "marker": "D"},
      "ORahilly_R1_Irish.xlsx":      {"colour": "#2ee828", "marker": "s"},
      "ORahilly_R2_Irish.xlsx":      {"colour": "#89c96a", "marker": "o"},
      "ORahilly_Combined":           {"colour": "#74a435", "marker": "*"},
      # Other datasets
      "Kinsella.txt":                {"colour": "red",    "marker": "v"},
      "Kinsella_NoPretales.txt":     {"colour": "blue",   "marker": "P"},
      "Carson.txt":                  {"colour": "orange", "marker": "p"},}
      style = dataset_styles.get(Name, {"colour": "black", "marker": "x"})

      plt.step(N, emp, style['marker'], ms = 4, label = Name, alpha=0.6, color=style['colour']) #color='k'
      plt.plot(Input, fit, alpha=0.4, color=style['colour']) #label = result[1]+' '+Name
      plt.xscale('log')
      plt.yscale('log')
      plt.xlabel('$k$', color = 'k', fontsize = 14)
      plt.ylim(ymin=0.5*np.min(emp))
      if dist == 'PDF':
        plt.ylabel(r'$p_k$', fontsize = 14)
      if dist == 'CCDF':
        plt.ylabel(r'$P_k$', fontsize = 14)
      plt.title('MLE Forced Power Law Results  ' + r'$k_{\rm min} = $'  + str(result[0]), fontsize=14)
      plt.legend() #bbox_to_anchor=(1.05, 1), loc='upper left'

      if save == True:
        if saveloc == None:
          raise ValueError
          'save is True but no save location provided. Please enter a folder path to save plot to'
          #raises an error if save is true but no filepath is provided.
        else:
          plt.savefig(saveloc + Name + ' ' + dist, dpi=600)
      #plt.show()

    def bootstrap(G_list, result):
        params1 = []
        params2 = []
        while len(params1) < 1000:
            sample = np.array(np.random.choice(G_list, len(G_list), replace=True))
            opt = opt_single_dist(sample, result, result[0])
            if len(opt) == 1: # if it is a one parameter distribution
                if np.isnan(opt[0]) == False: # we only count non-nan parameter values
                    params1.append(opt[0])
            if len(opt) == 2:
                if np.isnan(opt[0]) == False & np.isnan(opt[1]) == False:
                    params1.append(opt[0])
                    params2.append(opt[1])
        if len(params2) == 0: # graph has one distribution
          parameters = [params1]
        else:
            parameters = [params1, params2]
        return parameters

    def summary_stats(Name, result, params):
        means = []
        devs = []
        perc1 = []
        perc2 = []
        for i in params:
            means.append(np.round(np.mean(i), 2))
            devs.append(np.round(np.std(i), 2))
            perc1.append(np.round(np.percentile(i, 2.5), 2))
            perc2.append(np.round(np.percentile(i, 97.5), 2))
        row = [Name, result[1], result[0], result[2][0][0], means[0], devs[0], perc1[0], perc2[0]]
        if len(result[2][0]) == 2:
            row.extend([result[2][0][1], means[1], devs[1], perc1[1], perc2[1]])
        return row

    def fit(Name, G, k_min:int=1, vt=None, plot_type='auto', save=False, saveloc=None, IC='AIC'):
        X = degree_list(G) # Get the degree list
        if len(X) < 2500:
            if vt == None:
                vt = 2 # for small datasets we need fewer consecutive votes to
                      # determine the distribution
            if plot_type == 'auto':
                plot_type = 'ccdf' # for small datasets we do not want to display the PDF
        else:
            if vt == None:
                vt = 3
            if plot_type == 'auto':
                plot_type = 'both' # for larger datasets we display both PDF and
                                  # CCDF
        result = MLE(X, k_min, vt, IC) # Perform the MLE using kmin=1 as default
        N, P, p = empirical(X)	# Get the unique degree list, empirical CCDF and
                                # PDF values
        Input = np.arange(result[0],np.amax(X)+1) # generate complete list of integers
                                                  # kmin to kmax
        if plot_type == 'both' or plot_type == 'pdf':
            pdf = PDF(result, Input, N, P)
            plotting(N,Input, pdf, result, p, 'PDF', Name, save, saveloc)
        if plot_type == 'both' or plot_type == 'ccdf':
            ccdf = CCDF(result, Input, N, P)
            plotting(N,Input, ccdf, result, P, 'CCDF', Name, save, saveloc)
        print('For k greater than or equal to', result[0], 'the degree distribution follows a',
              result[1], 'distribution with parameters', np.round(result[2][0],2))
        return result

    plt.figure(figsize=(10,10))
    result = fit('',G,k_min=kmin,plot_type='ccdf')
    plt.show()

"""### Centrality Measures
- Ranking Nodes in terms of Degree
- Ranking Nodes in terms of Betweenness
"""

def centrality_measures(G,tables=True):
  '''
  === Summary ===
  Ranks characters in terms of their degree and betweenness

  === Inputs ===
  G : NetworkX graph
  tables (optional) : If True, prints table of the top characters

  === Outputs ===
  k_seq : List of characters in order of degree
  bet_seq : List of characters in order of betweenness
  '''
  # Rank nodes in order of degree
  k_seq = sorted(G.degree, key=lambda x: x[1], reverse=True)[:10]

  # Rank nodes in order of betweenness
  bet = nx.betweenness_centrality(G)
  bet_seq = {k: v for k, v in sorted(bet.items(), key=lambda item: item[1], reverse=True)[:10]}

  # Table of Top Characters in terms of Degree
  table_degree = PrettyTable()
  table_degree.field_names = ['Rank', 'Character', 'Degree']
  for i in range(10):
    table_degree.add_row([i+1, k_seq[i][0], k_seq[i][1]])

  # Table of Top Characters in terms of Betweenness
  table_betweenness = PrettyTable()
  table_betweenness.field_names = ['Rank', 'Character', 'Betweenness']
  for i in range(10):
    table_betweenness.add_row([i+1, list(bet_seq)[i], round(list(bet_seq.values())[i],2)])

  # Print tables
  if tables:
    print(table_degree)
    print(table_betweenness)

  return k_seq,bet_seq

"""### Spearman's Rank & Jaccard Index for Character Lists"""

import scipy as sc
def jaccard(l1, l2):
    s1, s2 = set(l1), set(l2)
    return len(s1.intersection(s2))/len(s1.union(s2))
def spearman(l1,l2):
    result = sc.stats.spearmanr(l1,l2)
    return result

"""### Random Simulations
- Reduces larger network by randomly removing nodes
"""

def random_simulations(G1,G2, table=True):
  '''
  === Summary ===
  Randomly removes nodes from larger network so the networks have equal N
  Calculates the properties of the reduced networks
  Compares the properties of the reduced networks to the original networks

  === Inputs ===
  G1 : NetworkX graph
  G2 : NetworkX graph
  table (optional) : If True, prints table of the top characters

  === Ouputs ===
  Prints table of values
  '''

  # Randomly remove nodes using random.sample() so both networks have equal N
  def remove_nodes(G,N_min):
    remove = random.sample(list(G.nodes()),len(G.nodes())-N_min) # Randomly selects nodes to be removed
    G_new = G.copy() #copy graph
    G_new.remove_nodes_from(remove) # Remove all the nodes selected in the 'remove' list
    return G_new

  # Ensure G1 is smaller
  if len(G1) > len(G2):
    G1, G2 = G2, G1  # Swap G1 and G2 if G1 is larger
  size = len(G1)

  # Pre Truncation Quantities
  N1,L1,k_avg1,k_delta1,r1,C1,Crand1,l_avg1,lrand1 = network_quantities(G1,table=False)
  N2,L2,k_avg2,k_delta2,r2,C2,Crand2,l_avg2,lrand2 = network_quantities(G2,table=False)

  # Truncation Comparison
  Ns = []
  Ls = []
  kavgs = []
  kdelts = []
  rs = []
  Cs = []
  Crands = []
  lavgs = []
  lrands = []
  IM_post = []
  #GED_post = []
  NMI_post = []
  KS_post = []

  # Generate random subgraphs of G2 and calculate the comparison metrics
  for i in range(100):
    G_new = remove_nodes(G2,size) # Create a truncated subgraph of the larger network (G2)
    N,L,k_avg,k_delta,r,C,Crand,l_avg,lrand = network_quantities(G_new,table=False) # Calculate the properties of the truncated network
    Ns.append(N)
    Ls.append(L)
    kavgs.append(k_avg)
    kdelts.append(k_delta)
    rs.append(r)
    Cs.append(C)
    Crands.append(Crand)
    lavgs.append(l_avg)
    lrands.append(lrand)
    IM_post.append(graph_matching(G1,G_new)[0])
    NMI_post.append(graph_matching(G1,G_new)[2])
    #GED_post.append(graph_matching(G1,G_new)[1])
    KS_post.append(scipy.stats.ks_2samp(G1,G_new))

  # Calculate average properties and metrics for the truncated subgraphs
  Navg = np.mean(Ns)
  Lavg = np.mean(Ls)
  kavg_new = np.mean(kavgs)
  kdelta_new = np.mean(kdelts)
  ravg = np.mean(rs)
  Cavg = np.mean(Cs)
  Crand_new = np.mean(Crands)
  l_avg_new = np.mean(lavgs)
  lrand_new = np.mean(lrands)
  IM_avg = np.mean(IM_post)
  #GED_avg = np.mean(GED_post)
  NMI_avg = np.mean(NMI_post)
  KS_stat_avg = np.mean(KS_post)

  if table:
    network_properties = PrettyTable()
    network_properties.field_names = ['Property','Smaller Network','Larger Network',              'Reduced Larger Network', 'Larger v Reduced Large', 'Smaller v Reduced Large']
    network_properties.add_row(['N',      N1,               N2,                                       round(Navg,0),        abs(round((N2 - Navg), 0)), abs(round((N1 - Navg), 0))])
    network_properties.add_row(['L',      L1,               L2,                                       round(Lavg,0),        abs(round((L2 - Lavg), 0)), abs(round((L1 - Lavg), 0))])
    network_properties.add_row(['<k>',    round(k_avg1,2),  round(k_avg2,2),                          round(kavg_new,2),    abs(round((k_avg2 - kavg_new), 2)), abs(round((k_avg1 - kavg_new), 2))])
    network_properties.add_row(['<k^2>',  round(k_delta1,2),round(k_delta2,2),                        round(kdelta_new,2),  abs(round((k_delta2 - kdelta_new), 2)), abs(round((k_delta1 - kdelta_new), 2))])
    network_properties.add_row(['r',      round(r1,2),      round(r2,2),                              round(ravg,2),        abs(round((r2 - ravg),2)), abs(round((r1 - ravg),2))])
    network_properties.add_row(['C',      round(C1,2),      round(C2,2),                              round(Cavg,2),        abs(round((C2 - Cavg),2)), abs(round((C1 - Cavg),2))])
    network_properties.add_row(['C_rand', round(Crand1, 2), round(Crand2, 2),                         round(Crand_new,2),   abs(round((Crand2 - Crand_new),2)), abs(round((Crand1 - Crand_new),2))])
    network_properties.add_row(['l_avg',  round(l_avg1,2),  round(l_avg2,2),                          round(l_avg_new,2),   abs(round((l_avg2 - l_avg_new),2)), abs(round((l_avg1 - l_avg_new),2))])
    network_properties.add_row(['l_rand', round(lrand1,2),  round(lrand2,2),                          round(lrand_new,2),   abs(round((lrand2 - lrand_new),2)), abs(round((lrand2 - lrand_new),2))])
    network_properties.add_row(['IM',     ' ',              round(graph_matching(G1,G2)[0],2),        round(IM_avg,2),      abs(round((graph_matching(G1,G2)[0] - IM_avg),2)), ' ' ])
    network_properties.add_row(['GED',    ' ',              round(graph_matching(G1,G2)[1],2),        ' ', ' ', ' '])
    network_properties.add_row(['NMI',    ' ',              round(graph_matching(G1,G2)[2],2),        round(NMI_avg,2),     abs(round((graph_matching(G1,G2)[2] - NMI_avg),2)), ' '  ])
    network_properties.add_row(['KS',     ' ',              round(scipy.stats.ks_2samp(G1,G2)[0],2),  round(KS_stat_avg,2), abs(round((scipy.stats.ks_2samp(G1,G2)[0] - KS_stat_avg),2)), ' '])
    print(network_properties)

    return

"""### Execute Code
- Run all of the functions mentioned above to create and analyse each networks
"""

def network_analysis(input_data, quantities=False, plot=False, matching=False, MLE_degdist=False, forcedPL_degdist=False,centrality=False,randomsims=False):
  '''
  === Summary ===
  Perform multiple types of network analysis on input datasets.
  This function loads networks from input data files, optionally performs:
  - Structural property calculations
  - Network visualisation
  - Graph matching metrics (IM, GED, NMI)
  - Degree distribution analysis (MLE or forced power-law)
  - Centrality measure comparison (degree & betweenness)
  - Random node removal simulation comparisons

  === Inputs ===
  input_data : List of filenames or paths to input network data.
  quantities : If True, calculate and display structural properties for each network.
  plot : If True, generate network plots.
  matching : If True, compute graph matching metrics (IM, GED, NMI)
  MLE_degdist : If True, performs MLE-based degree distribution analysis
  forcedPL_degdist : If True, performs forced power-law degree distribution analysis.
  centrality : If True, compares centrality measures across networks.
  randomsims : If True, run random simulation comparisons after 'ORahilly_Combined'.

  === Outputs ===
  Results are printed and/or plotted directly.
  '''

  # ===-----------------===
  # === Create Networks ===
  # ===-----------------===
  graphs = {} # dictionary for filenames and their associated graphs
  # Iterate through each file and create network
  for i in input_data:
    G = create_network(i)
    graphs[i] = G # store the graphs in a dictionary under it's filename

  # Add OR-comb network for networks without groups
  if 'ORahilly_R1_Irish.xlsx' in graphs and 'ORahilly_R2_Irish.xlsx' in graphs:
    g1 = graphs['ORahilly_R1_Irish.xlsx']
    g2 = graphs['ORahilly_R2_Irish.xlsx']
    orcomb = nx.compose(g1, g2)
    # New dict to insert OR-comb graph after ORII
    new_graphs = {}
    for key in graphs:
        new_graphs[key] = graphs[key]
        # Insert OR-comb after ORII
        if key == 'ORahilly_R2_Irish.xlsx':
            new_graphs['ORahilly_Combined'] = orcomb  # new key name for the composed graph
    graphs = new_graphs

  # Add OR-comb network for networks with groups
  if 'ORahilly_R1_Irish_Groups.xlsx' in graphs and 'ORahilly_R2_Irish_Groups.xlsx' in graphs:
    g1 = graphs['ORahilly_R1_Irish_Groups.xlsx']
    g2 = graphs['ORahilly_R2_Irish_Groups.xlsx']
    orcomb = nx.compose(g1, g2)
    # New dict to insert OR-comb graph after ORII
    new_graphs = {}
    for key in graphs:
        new_graphs[key] = graphs[key]
        # Insert OR-comb after ORII
        if key == 'ORahilly_R2_Irish_Groups.xlsx':
            new_graphs['ORahilly_Combined'] = orcomb  # new key name for the composed graph
    graphs = new_graphs


  # ===---------------------------===
  # === Network Global Properties ===
  # ===---------------------------===
  if quantities:
    # Iterate through all graphs and calculate their network global properties
    for i in graphs:
      print(i.rsplit('.', 1)[0])
      network_quantities(graphs[i])

  # ===-------------------===
  # === Plotting Networks ===
  # ===-------------------===
  if plot:
    # Iterate through the networks and plot them
    for filename, G in graphs.items():
      networkname = filename.rsplit('.', 1)[0]
      plot_network(G, save_path=f"{networkname}.png") # save network plots


  # ===----------------===
  # === Graph Matching ===
  # ===----------------===
  if matching:
    # Initialise empty dictionaries for the graph matching metrics
    IM_matrix = {}
    GED_matrix = {}
    NMI_matrix = {}
    file_list = list(graphs.keys()) # list of filenames

    # Iterate through each network in file_list
    for i, x in enumerate(file_list): #i=index,x=filename
      IM_matrix[x] = {}
      GED_matrix[x] = {}
      NMI_matrix[x] = {}

      for j, y in enumerate(file_list):
        if j < i:  # matrix is symmetric -> copy results to symmetric cell
          IM_matrix[x][y] = IM_matrix[y][x]
          GED_matrix[x][y] = GED_matrix[y][x]
          NMI_matrix[x][y] = NMI_matrix[y][x]
        else:
          # Calculte distance measures
          IM, GED, NMI_val = graph_matching(graphs[x], graphs[y])
          IM_matrix[x][y] = IM
          GED_matrix[x][y] = GED
          NMI_matrix[x][y] = NMI_val

    # Print the tables
    matrix_table(IM_matrix, "Ipsen-Mikhailov Distance",file_list)
    matrix_table(GED_matrix, "Graph Edit Distance",file_list)
    matrix_table(NMI_matrix, "Normalized Mutual Information",file_list)


  # ===----------------------------------------------===
  # === Degree Distribution Analysis - MLE & KS Test ===
  # ===----------------------------------------------===
  if MLE_degdist:
    KS_matrix = {}
    file_list = list(graphs.keys()) #list of filenames

    for i in graphs:
      print(i.rsplit('.', 1)[0])
      degree_distributions(graphs[i],kmin=2)

    for i,x in enumerate(file_list):
      KS_matrix[x] = {}
      for j,y in enumerate(file_list):
        if j < i:
          KS_matrix[x][y] = KS_matrix[y][x]
        elif j==1:
          KS_matrix[x][y] = 0.00
        else:
          # KS from the graphs
          KS = round(scipy.stats.ks_2samp(graphs[x],graphs[y])[0],2)
          KS_matrix[x][y] = KS
    matrix_table(KS_matrix, "Kolmogorov-Smirnov Test",file_list)


  # ===----------------------------------------------------===
  # === Degree Distribution Analysis - Forced PL & KS Test ===
  # ===----------------------------------------------------===
  # Degree distribution analysis - Forced PL
  if forcedPL_degdist:
    KS_matrix = {}
    file_list = list(graphs.keys()) #list og filenames
    plt.figure(figsize=(10,10))

    for i in graphs:
      print(i.rsplit('.', 1)[0])
      degree_distributions(graphs[i],kmin=3)
    plt.show()

    for i,x in enumerate(file_list):
      KS_matrix[x] = {}
      for j,y in enumerate(file_list):
        if j < i:
          KS_matrix[x][y] = KS_matrix[y][x]
        elif j==1:
          KS_matrix[x][y] = 0.00
        else:
          # Extract numeric sequences from the graphs
          KS = round(scipy.stats.ks_2samp(graphs[x],graphs[y])[0],2)
          KS_matrix[x][y] = KS
    matrix_table(KS_matrix, "Kolmogorov-Smirnov Test",file_list)


  # ===-----------------------------===
  # === Centrality Measure Analysis ===
  # ===-----------------------------===
  if centrality:
    # Initialise results matrices
    Spearman_matrix_k = {}
    Jaccard_matrix_k = {}
    Spearman_matrix_b = {}
    Jaccard_matrix_b = {}
    file_list = list(graphs.keys()) #list of filenames

    # Prints the character rankings
    for i in graphs:
      print(i.rsplit('.', 1)[0])
      centrality_measures(graphs[i])

    # Compare centrality rankings
    for i,x in enumerate(file_list):
      Spearman_matrix_k[x] = {}
      Jaccard_matrix_k[x] = {}
      Spearman_matrix_b[x] = {}
      Jaccard_matrix_b[x] = {}

      for j,y in enumerate(file_list):
        if j < i:
          # Copy symmetric values
          Spearman_matrix_k[x][y] = Spearman_matrix_k[y][x]
          Jaccard_matrix_k[x][y] = Jaccard_matrix_k[y][x]
          Spearman_matrix_b[x][y] = Spearman_matrix_b[y][x]
          Jaccard_matrix_b[x][y] = Jaccard_matrix_b[y][x]

        elif j==1:
          Spearman_matrix_k[x][y] = 1.00
          Jaccard_matrix_k[x][y] = 1.00
          Spearman_matrix_b[x][y] = 1.00
          Jaccard_matrix_b[x][y] = 1.00

        else:
          # Extract degree and betweenness centrality measures for x
          kx,bx = centrality_measures(graphs[x],tables=False)
          kxnames = [x for x,_ in kx]
          bxnames = list(bx.keys())

          # Extract for y
          ky,by = centrality_measures(graphs[y],tables=False)
          kynames = [x for x,_ in ky]
          bynames = list(by.keys())

          # Compare the rankings using Spearman's rank and Jaccard Index
          Spearman_k = spearman(kxnames,kynames)
          Jaccard_k = jaccard(kxnames,kynames)
          Spearman_matrix_k[x][y] = Spearman_k[0]
          Jaccard_matrix_k[x][y] = Jaccard_k

          Spearman_b = spearman(bxnames,bynames)
          Jaccard_b = jaccard(bxnames,bynames)
          Spearman_matrix_b[x][y] = Spearman_b[0]
          Jaccard_matrix_b[x][y] = Jaccard_b

    # Print the tables
    print('Degree')
    matrix_table(Spearman_matrix_k, "Spearmans Rank",file_list)
    matrix_table(Jaccard_matrix_k, "Jaccard Index",file_list)
    print('\n','Betweenness')
    matrix_table(Spearman_matrix_b, "Spearmans Rank",file_list)
    matrix_table(Jaccard_matrix_b, "Jaccard Index",file_list)


  # ===-----------------------------===
  # === Random Simulations Analysis ===
  # ===-----------------------------===
  if randomsims:
    file_list = list(graphs.keys())
    # Loop through each graph
    for i, x in enumerate(file_list):
        print(f"Comparing {x.rsplit('.',1)[0]}")

        # Compare i to j
        for j, y in enumerate(file_list):
            if i != j:
                print(f"  vs {y.rsplit('.',1)[0]}")
                random_simulations(graphs[x], graphs[y],table=True)

network_analysis(nogroups,plot=True)